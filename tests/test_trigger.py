#!/usr/bin/env python3
"""test_trigger.py — chiguo_trigger 触发器引擎单元测试（v2 sigmoid 加权随机）"""

import os
import random
import sys
import tempfile
import tomllib
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

CST = timezone(timedelta(hours=8))

from chiguo_state import ChiguoState
from chiguo_trigger import evaluate_triggers


def _make_state(tmp: str, now: datetime, **overrides) -> ChiguoState:
    """真实 toml 配置 + 临时目录锚定；mem0 指向不存在路径 + 显式禁用 → available=False（确定性）"""
    cfg_path = Path(tmp) / "chiguo_proactive.toml"
    cfg_path.write_text(Path("chiguo_proactive.toml").read_text())
    with open(cfg_path, "rb") as f:
        cfg = tomllib.load(f)
    cfg["_base_dir"] = str(tmp)
    cfg["memory"]["mem0_qdrant_path"] = str(Path(tmp) / "no_qdrant")
    cfg["memory"]["mem0_history_db"] = str(Path(tmp) / "no_history.db")
    os.environ["CHIGUO_MEM0_DISABLED"] = "1"
    s = ChiguoState(cfg)
    # 默认 10h 前的用户消息 → silent≈6h（睡眠窗口 0-8 抵消 4h），介于 (2,48)
    s.cooldown.last_user_message_at = (now - timedelta(hours=10)).isoformat()
    s.cooldown.current_date = now.strftime("%Y-%m-%d")
    for k, v in overrides.items():
        if hasattr(s.emotion, k):
            setattr(s.emotion, k, v)
        elif hasattr(s.cooldown, k):
            setattr(s.cooldown, k, v)
    return s


def _run_seeds(s: ChiguoState, now: datetime, n: int = 200, seed0: int = 1000) -> dict:
    """固定种子序列逐一运行 evaluate_triggers → type 计数（确定性，不依赖模块级 RNG）"""
    counts: dict[str, int] = {}
    for i in range(n):
        random.seed(seed0 + i)
        t = evaluate_triggers(s, now)
        key = t.type if t else "None"
        counts[key] = counts.get(key, 0) + 1
    return counts


def _count_must_send(s: ChiguoState, now: datetime, n: int = 200, seed0: int = 1000) -> int:
    """固定种子序列统计 must_send 标记次数（A4 高段边界断言用）"""
    ms = 0
    for i in range(n):
        random.seed(seed0 + i)
        t = evaluate_triggers(s, now)
        if t and t.data.get("must_send"):
            ms += 1
    return ms


# ═══════════════════════════════════════════════════════════
# 孤独三级 softmax 归一化竞争
# ═══════════════════════════════════════════════════════════

def test_low_loneliness_never_fires_lonely_or_anxiety():
    """低孤独(15)+默认焦虑(40)：归一化后 lonely/anxiety 均非候选（实测 200 种子 0 次）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now)
        counts = _run_seeds(s, now)
        lonely = sum(v for k, v in counts.items() if k.startswith("lonely_"))
        assert lonely == 0, f"lonely must not fire at loneliness=15, got {counts}"
        assert counts.get("anxiety", 0) == 0, f"anxiety must not fire at anxiety=40, got {counts}"
        # 14:00 无 ritual 候选 + 无记忆 → 唯一候选为 playful（实测 200/200）
        assert counts.get("playful", 0) == 200, f"playful should be sole candidate, got {counts}"
    print("  OK test_low_loneliness_never_fires_lonely_or_anxiety")


def test_no_candidates_returns_none():
    """整体不触发：低孤独 + 低元气（playful 门槛 energy>70 不满足）→ 候选集空 → None"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        counts = _run_seeds(s, now, n=200)
        assert counts.get("None", 0) == 200, f"expected all None, got {counts}"
    print("  OK test_no_candidates_returns_none")


def test_lonely_softmax_competition_at_high_loneliness():
    """高孤独(75)：三级 softmax 归一化互斥竞争（实测 200 种子：low 92 / mid 90 / high 18）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, loneliness=75, energy=40)
        counts = _run_seeds(s, now, n=200)
        lonely = sum(v for k, v in counts.items() if k.startswith("lonely_"))
        assert lonely == 200, f"all triggers should be lonely at loneliness=75, got {counts}"
        # 种子固定为确定性序列 → 三级均非零（实测 200 种子：low 92 / mid 90 / high 18）
        assert counts.get("lonely_low", 0) > 0, f"softmax must spread to low level, got {counts}"
        assert counts.get("lonely_mid", 0) > 0, f"softmax must spread to mid level, got {counts}"
        assert counts.get("lonely_high", 0) > 0, f"softmax must spread to high level, got {counts}"
        assert counts.get("anxiety", 0) == 0, f"anxiety must not fire at anxiety=40, got {counts}"
    print("  OK test_lonely_softmax_competition_at_high_loneliness")


def test_rate_factor_boosts_lonely_at_low_loneliness():
    """暴涨变化率：孤独=15 但 rate=10 → rate_factor=3.55 → lonely_low 进入候选（高概率触发）
    v10 (#73 A4): 孤独 15 基础权重小 → activation 在 min_activation(0.08) 边缘，
    ×1.2 抖动下限偶尔跌入低段沉默 → 断言放宽为 ≥150/200 lonely_low，其余 None"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        s.emotion.loneliness_rate = 10.0
        counts = _run_seeds(s, now, n=200)
        low = counts.get("lonely_low", 0)
        assert low >= 150, f"rate 暴涨应高概率触发 lonely_low, got {counts}"
        assert low + counts.get("None", 0) == 200, f"仅 lonely_low/None 允许, got {counts}"
    print("  OK test_rate_factor_boosts_lonely_at_low_loneliness")


# ═══════════════════════════════════════════════════════════
def test_anxiety_high_fires():
    """anxiety=80：w≈0.65 → 高概率触发（实测 100 种子 88 次 anxiety + 12 次 playful）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, anxiety=80)
        counts = _run_seeds(s, now, n=100)
        assert counts.get("anxiety", 0) >= 50, f"anxiety=80 should fire often, got {counts}"
        assert counts.get("None", 0) == 0, f"anxiety candidate should always exist, got {counts}"
    print("  OK test_anxiety_high_fires")


def test_b2_anxiety_extreme_reaches_must_send():
    """B2 (#137): A4 must_send 标定矛盾修复 —— 修复前 w_anx 被 "不触发基线" softmax
    钳在 max≈0.664 < must_send_activation(0.75)，anxiety 单源权重永远到不了高段必发
    （空闲×1.2 下 anx=70 仅 0.74 <0.75 不触发，红线）。
    修复后 raw_anx→1 时 w_anx→≈1：空闲高焦虑 70 → 必发标记；极高焦虑 90(raw≈0.98)
    → 基础权重≈0.99 ≥0.75，端到端 100% 标记 must_send；
    中低焦虑(40) 归一化仍 ≈0.187<0.3 不成候选 → 不触发 must_send（恒等）。"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)  # 空闲 ×1.2
        # 高焦虑 70（= anxiety_block_threshold）+ 低孤独 + 低元气（关 playful 噪声）
        # → anxiety 为唯一情绪候选。修复前 w_old≈0.618×1.2=0.741<0.75 不标记（红线）；
        # 修复后 w_new≈0.894×1.2=1.07 ≥0.75 → 必发。
        s70 = _make_state(td, now, anxiety=70, loneliness=15, energy=40)
        ms70 = _count_must_send(s70, now, n=50)
        assert ms70 == 50, f"高焦虑 70 空闲应必发 must_send, got {ms70}/50"
        # 极高焦虑 90：raw_anx≈0.98 → 修复后基础 w_anx≈0.99 ≥0.75，端到端 100% 标记
        s90 = _make_state(td, now, anxiety=90, loneliness=15, energy=40)
        ms90 = _count_must_send(s90, now, n=50)
        assert ms90 == 50, f"极高焦虑 90 应 100% must_send, got {ms90}/50"
        random.seed(99)
        t = evaluate_triggers(s90, now)
        assert t is not None and t.type == "anxiety", f"极高焦虑应选中 anxiety, got {t}"
        assert t.data.get("must_send") is True, f"must_send 标记缺失, got {t.data}"
        # 中低焦虑(40) 对照：归一化 <0.3 不成候选 → 不得 must_send
        s40 = _make_state(td, now, anxiety=40, loneliness=15, energy=40)
        ms40 = _count_must_send(s40, now, n=50)
        assert ms40 == 0, f"中低焦虑不得 must_send, got {ms40}/50"
    print("  OK test_b2_anxiety_extreme_reaches_must_send")


# ═══════════════════════════════════════════════════════════
# 时间窗口 ritual 触发
# ═══════════════════════════════════════════════════════════

def test_morning_window_probability():
    """早安窗口 09:00（8-10h）：10% 概率触发（实测 200 种子 18 次）；morning_sent 后永不触发"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 9, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        counts = _run_seeds(s, now, n=200)
        morning = counts.get("morning", 0)
        assert 8 <= morning <= 40, f"morning ~20/200 expected, got {counts}"
        assert set(counts) <= {"morning", "None"}, f"only morning/None allowed, got {counts}"
        # morning_sent → 门控关闭
        s.cooldown.morning_sent = True
        counts2 = _run_seeds(s, now, n=50)
        assert counts2.get("morning", 0) == 0, f"morning_sent must suppress, got {counts2}"
    print("  OK test_morning_window_probability")


def test_night_window_probability():
    """晚安窗口 20:00（20-21h）：12% 概率触发（实测 200 种子 24 次）；night_sent 后永不触发"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 20, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        counts = _run_seeds(s, now, n=200)
        night = counts.get("night", 0)
        assert 10 <= night <= 45, f"night ~24/200 expected, got {counts}"
        assert set(counts) <= {"night", "None"}, f"only night/None allowed, got {counts}"
        s.cooldown.night_sent = True
        counts2 = _run_seeds(s, now, n=50)
        assert counts2.get("night", 0) == 0, f"night_sent must suppress, got {counts2}"
    print("  OK test_night_window_probability")


def test_meal_window_probability():
    """饭点 12:00（11/12/17/18/19h）：5% 概率触发（实测 300 种子 12 次）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 12, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        counts = _run_seeds(s, now, n=300)
        meal = counts.get("meal", 0)
        assert 4 <= meal <= 30, f"meal ~15/300 expected, got {counts}"
        assert set(counts) <= {"meal", "None"}, f"only meal/None allowed, got {counts}"
    print("  OK test_meal_window_probability")


# ═══════════════════════════════════════════════════════════
# playful / reflect / longing 候选
# ═══════════════════════════════════════════════════════════

def test_playful_requires_high_energy_and_silence():
    """playful 条件：energy>70 且 2<silent_h<48 且空闲（实测 energy=85/silent≈6h → 200/200）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now)
        counts = _run_seeds(s, now, n=200)
        assert counts.get("playful", 0) == 200, f"playful should be sole candidate, got {counts}"
        # energy=40 → 门槛不满足 → 不触发
        s2 = _make_state(td, now, energy=40)
        counts2 = _run_seeds(s2, now, n=100)
        assert counts2.get("playful", 0) == 0, f"energy=40 must not fire playful, got {counts2}"
        # silent_h>48 → 门槛不满足 → 不触发（无其他候选 → 全 None）
        s3 = _make_state(td, now, energy=85)
        s3.cooldown.last_user_message_at = (now - timedelta(hours=100)).isoformat()
        counts3 = _run_seeds(s3, now, n=100)
        assert counts3.get("playful", 0) == 0, f"silent>48 must not fire playful, got {counts3}"
        assert counts3.get("None", 0) == 100, f"expected all None, got {counts3}"
    print("  OK test_playful_requires_high_energy_and_silence")


def test_reflect_requires_high_affection_low_neuroticism():
    """reflect 条件：affection>70 & silent<2 & energy>60 & neuroticism<70 & 8% 概率门。
    v10 (#73 A4): 孤独 15 时 reflect 为唯一情绪候选且权重 ~0.02 < min_activation(0.08)
    → 低段沉默；改为孤独 25 提供 lonely_low 陪伴候选（activation 进入中段），
    reflect 候选 ~8% 门 × 竞争权重 ~12.5% → 实测 ~3/300，断言 1-15"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, affection=80, energy=85, loneliness=25)
        s.cooldown.last_user_message_at = (now - timedelta(minutes=30)).isoformat()
        counts = _run_seeds(s, now, n=300)
        reflect = counts.get("reflect", 0)
        assert 1 <= reflect <= 15, f"reflect ~3/300 expected, got {counts}"
        assert set(counts) <= {"reflect", "lonely_low"}, f"only reflect/lonely_low allowed, got {counts}"
        # neuroticism=80 → 门槛不满足 → 永不触发（全 lonely_low）
        s2 = _make_state(td, now, affection=80, energy=85, loneliness=25)
        s2.cooldown.last_user_message_at = (now - timedelta(minutes=30)).isoformat()
        s2.personality.neuroticism = 80.0
        counts2 = _run_seeds(s2, now, n=100)
        assert counts2.get("reflect", 0) == 0, f"neuroticism=80 must not fire reflect, got {counts2}"
    print("  OK test_reflect_requires_high_affection_low_neuroticism")


def test_longing_overflow_candidate():
    """longing 溢出：held>3 且 λ 累积 ≥ base×1.5 且焦虑不阻塞 → 加权候选（实测 100/100）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        s.cooldown.held_count = 5
        s.cooldown.accumulated_lambda = 0.5
        counts = _run_seeds(s, now, n=100)
        assert counts.get("longing", 0) == 100, f"expected all longing, got {counts}"
        random.seed(1234)
        t = evaluate_triggers(s, now)
        assert t.data.get("held_count") == 5, t.data
        assert t.data.get("accumulated_lambda") == 0.5, t.data
        # 负例：held=2 不溢出 → 无候选 → None
        s2 = _make_state(td, now, energy=40)
        s2.cooldown.held_count = 2
        s2.cooldown.accumulated_lambda = 0.5
        counts2 = _run_seeds(s2, now, n=50)
        assert counts2.get("longing", 0) == 0 and counts2.get("None", 0) == 50, f"got {counts2}"
    print("  OK test_longing_overflow_candidate")


def test_longing_overflow_zero_base_lambda_no_crash():
    """base_lambda=0 时 longing 分支不应 ZeroDivisionError（回归 B4）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        s.cooldown.held_count = 5
        s.cooldown.accumulated_lambda = 0.5
        s.config["poisson"]["base_lambda"] = 0
        t = evaluate_triggers(s, now)
        assert t is None or t.type != "longing", f"base_lambda=0 不应产出 longing, got {t}"
    print("  OK test_longing_overflow_zero_base_lambda_no_crash")


# ═══════════════════════════════════════════════════════════
# 特殊日期 ritual + 记忆触发（tz 防护）
# ═══════════════════════════════════════════════════════════

def test_special_date_ritual():
    """特殊日期(3c 起数据源 = anniversary_mgr 当天匹配,替代 toml special_dates):
    special 高权重候选(实测 100/100);非特殊日不触发"""
    with tempfile.TemporaryDirectory() as td:
        from pathlib import Path as _P
        import json as _json
        _P(td, "anniversaries.json").write_text(_json.dumps({"anniversaries": [
            {"id": "a1", "type": "anniversary", "name": "认识纪念日", "date": "11-03",
             "note": "", "created_at": "2026-01-01"}]}))
        now = datetime(2026, 11, 3, 14, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        counts = _run_seeds(s, now, n=100)
        assert counts.get("special", 0) == 100, f"expected all special, got {counts}"
        now2 = datetime(2026, 11, 4, 14, 0, tzinfo=CST)
        s2 = _make_state(td, now2, energy=40)
        counts2 = _run_seeds(s2, now2, n=50)
        assert counts2.get("special", 0) == 0, f"non-special day must not fire, got {counts2}"
    print("  OK test_special_date_ritual")


def test_memory_reminder_naive_tz_guard():
    """reminder 记忆：naive trigger_at 按 CST 补齐后 ±10min 窗口内触发（实测 50/50）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        naive = datetime(2026, 6, 15, 13, 55).isoformat()  # 无 tzinfo → 按 CST 解析
        s.memories = [{"type": "reminder", "trigger_at": naive, "content": "喝水"}]
        counts = _run_seeds(s, now, n=50)
        assert counts.get("memory", 0) == 50, f"expected all memory, got {counts}"
    print("  OK test_memory_reminder_naive_tz_guard")


def test_memory_reminder_garbage_at_does_not_crash():
    """垃圾 trigger_at 字符串 → 不崩、不触发（实测 30 种子全 None）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        s.memories = [{"type": "reminder", "trigger_at": "garbage", "content": "x"}]
        counts = _run_seeds(s, now, n=30)
        assert counts.get("None", 0) == 30, f"garbage trigger_at must be skipped, got {counts}"
    print("  OK test_memory_reminder_garbage_at_does_not_crash")


def test_memory_reminder_dedup_after_trigger():
    """reminder 去重（#79）：① mem 已标记 last_triggered_at（daemon 发送后写入）
    → 同进程不再触发；② trigger_at 在未来 → 窗口收紧（now >= t）不提前触发
    （旧 abs<600 逻辑会误触发未来 5 分钟的提醒）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        # ① 已触发过 → 去重
        s = _make_state(td, now, energy=40)
        s.memories = [{"type": "reminder", "trigger_at": "2026-06-15T13:55",
                       "content": "喝水", "last_triggered_at": "2026-06-15T14:00:00+08:00"}]
        counts = _run_seeds(s, now, n=50)
        assert counts.get("memory", 0) == 0, \
            f"已标记 last_triggered_at 不得重复触发, got {counts}"
        # ② 提前触发（trigger_at 在未来 5 分钟）→ 不触发
        s2 = _make_state(td, now, energy=40)
        s2.memories = [{"type": "reminder", "trigger_at": "2026-06-15T14:05", "content": "喝水"}]
        counts2 = _run_seeds(s2, now, n=30)
        assert counts2.get("memory", 0) == 0, \
            f"提前触发应被窗口收紧拦截（0 <= now-t < 600）, got {counts2}"
    print("  OK test_memory_reminder_dedup_after_trigger")


def test_memory_habit_window_probability():
    """habit 记忆：窗口小时命中 → 6% 概率触发（实测 300 种子 15 次）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        s.memories = [{"type": "habit", "trigger_window": [14], "content": "背单词"}]
        counts = _run_seeds(s, now, n=300)
        memory = counts.get("memory", 0)
        assert 2 <= memory <= 30, f"habit ~18/300 expected, got {counts}"
    print("  OK test_memory_habit_window_probability")


# ═══════════════════════════════════════════════════════════
# v10 (#73): A3 日程乘数+抖动 / A4 三段激活 / A6 repeat 阻尼
# ═══════════════════════════════════════════════════════════

def _inject_xlsx(tmp: str) -> None:
    """生成最小课表 fixture 到临时目录（与 test_integration.test_7 同手法，自包含）→ schedule 可用"""
    from openpyxl import Workbook
    dst = Path(tmp) / "data" / "xskb.xlsx"
    dst.parent.mkdir(parents=True, exist_ok=True)
    _wb = Workbook()
    _ws = _wb.active
    _ws.cell(row=5, column=1, value=1)
    _ws.cell(row=5, column=2, value="高等数学BII(理论)-刘洋【2-17周】尚行楼")
    _wb.save(str(dst))


def test_a3_schedule_multiplier_tiers():
    """A3 日程乘数三态（单元级确定性）：空闲 free_multiplier / 上课 0.3 / 半忙 0.6"""
    from chiguo_trigger import _schedule_multiplier
    with tempfile.TemporaryDirectory() as td:
        # 空闲：无课表 → schedule_status None → _is_free_time True → free_multiplier
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        assert _schedule_multiplier(s, now, 1.2) == 1.2, "空闲 → free_multiplier"
        # 半忙：静默窗口内（非空闲、非上课）→ 0.6
        now3 = datetime(2026, 6, 15, 3, 0, tzinfo=CST)
        s3 = _make_state(td, now3, energy=40)
        assert _schedule_multiplier(s3, now3, 1.2) == 0.6, "静默时段 → 半忙 0.6"
        # 上课：注入课表 + 周一 08:30 第 1 节 → in_class → 0.3
        _inject_xlsx(td)
        now1 = datetime(2026, 6, 15, 8, 30, tzinfo=CST)
        s1 = _make_state(td, now1, energy=40)
        sch = s1.schedule_status(now1)
        assert sch and sch.get("in_class"), f"前置条件:周一 08:30 应上课中, got {sch}"
        assert _schedule_multiplier(s1, now1, 1.2) == 0.3, "上课中 → 0.3"
    print("  OK test_a3_schedule_multiplier_tiers")


def test_a3_in_class_suppresses_emotion_only():
    """A3 黑盒：上课时情绪类权重 ×0.3 压缩 → 仪式类 memory 反超（仪式类豁免）。
    F-A5-01（#314 R9）：reminder 视为用户显式托付、准时优先——空闲高段（must_send）
    不再压制 reminder（高段豁免）；无 reminder 时仍情绪类必发。"""
    with tempfile.TemporaryDirectory() as td:
        _inject_xlsx(td)
        # 上课 08:30（周一第 1 节）+ reminder 记忆（±10min 窗口）
        now = datetime(2026, 6, 15, 8, 30, tzinfo=CST)
        s = _make_state(td, now, loneliness=75, energy=40)
        s.cooldown.morning_sent = True  # 关闭 morning 概率门，防仪式类噪声
        s.memories = [{"type": "reminder", "trigger_at": "2026-06-15T08:25", "content": "喝水"}]
        counts_in = _run_seeds(s, now, n=200)
        lonely_in = sum(v for k, v in counts_in.items() if k.startswith("lonely_"))
        assert counts_in.get("memory", 0) > lonely_in, \
            f"上课时仪式类应反超情绪类, got {counts_in}"
        # 空闲 14:00 同配置（无课表）→ 上空高段必发：
        # ① reminder 在窗口内 → 高段豁免，必须发出（F-A5-01）
        now2 = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s2 = _make_state(td, now2, loneliness=75, energy=40)
        s2.memories = [{"type": "reminder", "trigger_at": "2026-06-15T13:55", "content": "喝水"}]
        counts_free = _run_seeds(s2, now2, n=200)
        assert counts_free.get("memory", 0) > 0, \
            f"空闲高段 reminder 应豁免必发, got {counts_free}"
        # ② 无 reminder 时 → 情绪类必发（A4 must_send 不回归）
        s3 = _make_state(td, now2, loneliness=75, energy=40)
        counts_free2 = _run_seeds(s3, now2, n=200)
        lonely_free2 = sum(v for k, v in counts_free2.items() if k.startswith("lonely_"))
        assert lonely_free2 == 200, f"空闲无 reminder 应全发情绪类, got {counts_free2}"
    print("  OK test_a3_in_class_suppresses_emotion_only")


def test_a4_low_activation_silences_emotion_ritual_fires():
    """A4 低段：孤独 15 → 情绪候选截断 → activation=0 < min_activation(0.08)
    → 情绪类不参与竞争；仪式类 memory 照发（100/100）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        s.memories = [{"type": "reminder", "trigger_at": "2026-06-15T13:55", "content": "喝水"}]
        counts = _run_seeds(s, now, n=100)
        assert counts.get("memory", 0) == 100, f"低段仪式类应照发, got {counts}"
    print("  OK test_a4_low_activation_silences_emotion_ritual_fires")


def test_a4_must_send_high_activation():
    """A4 高段：孤独 75 + 特殊日 → activation ≥ must_send_activation(0.75)
    → 情绪类必选（special 退让 0 次），选中结果标记 must_send: true；
    中段（孤独 30）→ 加权竞争 special 占优且不标记"""
    with tempfile.TemporaryDirectory() as td:
        from pathlib import Path as _P
        import json as _json
        _P(td, "anniversaries.json").write_text(_json.dumps({"anniversaries": [
            {"id": "a1", "type": "anniversary", "name": "认识纪念日", "date": "11-03",
             "note": "", "created_at": "2026-01-01"}]}))
        now = datetime(2026, 11, 3, 14, 0, tzinfo=CST)
        s = _make_state(td, now, loneliness=75, energy=40)
        counts = _run_seeds(s, now, n=100)
        assert counts.get("special", 0) == 0, f"must_send 应压制 special, got {counts}"
        lonely = sum(v for k, v in counts.items() if k.startswith("lonely_"))
        assert lonely == 100, f"高段应必发情绪类, got {counts}"
        # 标记验证：单次调用情绪类 → must_send=True
        random.seed(4242)
        t = evaluate_triggers(s, now)
        assert t is not None and t.type.startswith("lonely_"), t
        assert t.data.get("must_send") is True, f"must_send 标记缺失, got {t.data}"
        # 中段对照：孤独 30 → activation<0.5 → 加权竞争，special 占优且无 must_send
        s2 = _make_state(td, now, loneliness=30, energy=40)
        counts2 = _run_seeds(s2, now, n=100)
        assert counts2.get("special", 0) > 50, f"中段 special 应占优, got {counts2}"
        assert sum(v for k, v in counts2.items() if k.startswith("lonely_")) > 0, \
            f"中段情绪类应仍有概率, got {counts2}"
        random.seed(4242)
        t2 = evaluate_triggers(s2, now)
        assert t2 is not None and not t2.data.get("must_send"), \
            f"中段不得标记 must_send, got {t2.type} {t2.data}"
    print("  OK test_a4_must_send_high_activation")


def test_a4_mid_band_no_must_send():
    """A4 边界（#79）：must_send_activation=0.75 下，孤独 31-45 与焦虑 45-58
    区间不触发 must_send（200 种子 × 全区间 0 次）。
    半忙环境（03:00 静默窗口，日程乘数 ×0.6）→ activation（维度族 max，抖动前
    确定性）严格 < 0.75；空闲(×1.2) 下孤独 45 族和≈0.70×1.2=0.84 才越过阈值必发，
    孤独≤40 空闲仍 <0.75（R7 修正：0.75 按单源标定，两股中低情绪叠加不再凑段）。"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 3, 0, tzinfo=CST)  # 静默窗口 0-8 → 半忙 ×0.6
        # 孤独 31-45（焦虑 40 默认；energy=40 关闭 playful 噪声）
        for lo in range(31, 46):
            s = _make_state(td, now, loneliness=lo, anxiety=40, energy=40)
            ms = _count_must_send(s, now)
            assert ms == 0, f"loneliness={lo}: 中段不得触发 must_send, got {ms}/200"
        # 焦虑 45-58（孤独 15 默认；45-52 归一化 <0.3 不成候选 → 低段 None）
        for anx in range(45, 59):
            s = _make_state(td, now, anxiety=anx, energy=40)
            ms = _count_must_send(s, now)
            assert ms == 0, f"anxiety={anx}: 中段不得触发 must_send, got {ms}/200"
    print("  OK test_a4_mid_band_no_must_send")


def test_a4_must_send_free_band_boundary_anchor():
    """F-A5-03 (#315 R13)：A4 高段标定锚定测试——修复测试空洞：原
    test_a4_mid_band_no_must_send 只在 ×0.6 半忙环境（03:00 静默窗）断言，未覆盖
    空闲（×1.2）环境的标定边界。审计 E1 实测：空闲孤独≥42 / 焦虑≥60 → 100%
    must_send（高段=0.75 按单源标定），孤独≤40 / 焦虑≤55 仍 0。
    现状行为即正确（锚定不红），加上半忙对照确保空洞补全不回归。"""
    with tempfile.TemporaryDirectory() as td:
        now_free = datetime(2026, 6, 15, 14, 0, tzinfo=CST)  # 空闲 ×1.2（非静默非上课）
        # 边界外：空闲孤独 40 不得 must_send
        s40 = _make_state(td, now_free, loneliness=40, anxiety=40, energy=40)
        assert _count_must_send(s40, now_free) == 0, \
            f"空闲孤独=40 不得 must_send, got {_count_must_send(s40, now_free)}/200"
        # 高段锚定：空闲孤独 42/44/45 → 100% must_send（审计 E1 边界精确一致）
        for lo in (42, 44, 45):
            s = _make_state(td, now_free, loneliness=lo, anxiety=40, energy=40)
            ms = _count_must_send(s, now_free)
            assert ms == 200, f"空闲孤独={lo} 应 100% must_send, got {ms}/200"
        # 焦虑单源标定：55 边界外 0；60 高段 100%
        s55 = _make_state(td, now_free, anxiety=55, loneliness=15, energy=40)
        assert _count_must_send(s55, now_free) == 0, \
            f"空闲焦虑=55 不得 must_send, got {_count_must_send(s55, now_free)}/200"
        s60 = _make_state(td, now_free, anxiety=60, loneliness=15, energy=40)
        ms60 = _count_must_send(s60, now_free)
        assert ms60 == 200, f"空闲焦虑=60 应 100% must_send, got {ms60}/200"
        # 空洞补全对照：半忙 ×0.6（03:00 静默窗）孤独 45 → 0（不回归现有 test_a4_mid_band）
        now_half = datetime(2026, 6, 15, 3, 0, tzinfo=CST)
        sh = _make_state(td, now_half, loneliness=45, anxiety=40, energy=40)
        msh = _count_must_send(sh, now_half)
        assert msh == 0, f"半忙孤独=45 不得 must_send, got {msh}/200"
    print("  OK test_a4_must_send_free_band_boundary_anchor")


# ═══════════════════════════════════════════════════════════
# F-A5-01（#314 R9）: 高段必发只从情绪候选选 → reminder 一次性记忆确定性丢失
# 三机制中的①（高段压制）：只能从情绪候选选，ritual/memory 候选退让。
# 修复：reminder/MEMORY 候选在窗口内先于情绪高段分支处理（用户决策：提醒准时优先）。
#═══════════════════════════════════════════════════════════════

def _count_reminder_high_band(seed0=42):
    """高段（空闲×1.2，孤独45 → activation≥0.75）300 种子 reminder 触发统计。
    同审计 E1 复现构造：孤独45 + 空闲（非上课非静默）→ 情绪类必发压制。"""
    with tempfile.TemporaryDirectory() as td:
        # 14:00 非静默窗(0-8)外、无课表 → 空闲 ×1.2 → 孤独 45 踏过高段必发阈值
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, loneliness=45, energy=40)
        s.memories = [{"type": "reminder", "trigger_at": "2026-06-15T13:55", "content": "喝水"}]
        counts = _run_seeds(s, now, n=300, seed0=seed0)
        return counts


def test_fa5_reminder_triggered_in_high_band():
    """F-A5-01 ①：reminder 窗口内（触发时刻 ±30min）在常态高段（空闲×1.2 孤独45）
    必须被选中触发——修复前红：0/300（高段只从情绪候选选压制 memory）。

    用户决策「提醒准时优先」：reminder 视为显式托付，高段必须发出（豁免
    '只从情绪候选选'压制）。此断言为回归线：任何种子序列下 reminder 命中 >0。"""
    counts = _count_reminder_high_band()
    # 回归断言：reminder 以 MEMORY 类型候选触发（evaluate_triggers 返回
    # Trigger(type=MEMORY)，counts 键为 "memory"）
    memory = counts.get("memory", 0)
    assert memory > 0, f"F-A5-01: 高段窗口内 reminder 必须被触发, got {counts}"


def test_fa5_reminder_still_fires_low_band_baseline():
    """F-A5-01 低段对照：孤独 15（低段）reminder 仍可发（审计基线 285-300/300）。
    本回归线守卫「修复不破坏低段正常触发」——断言宽松上界保持可发（≥ 270/300）。"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        # 默认 energy=85 → playful 参与竞争（与审计 285/300 同构；此处 280/300）
        s = _make_state(td, now, loneliness=15)
        s.memories = [{"type": "reminder", "trigger_at": "2026-06-15T13:55", "content": "喝水"}]
        counts = _run_seeds(s, now, n=300, seed0=42)
        memory = counts.get("memory", 0)
        assert memory >= 270, \
            f"低段 reminder 不得回归（基线 285/300）, got {counts}"


def test_a4_must_send_preserved_across_safety_downgrade():
    """A4×安全阀组合：高孤独 must_send 选中 lonely_high 后被 safety≥1 降级为
    lonely_mid/soft 时，data 必须继承（must_send 标记不得丢失）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, loneliness=75, energy=40)
        # safety=1：24h 内有崩溃记录
        s.cooldown.last_crash_at = now.isoformat()
        s.cooldown.crash_timestamps = [now.isoformat()]
        seen_downgrade = 0
        for i in range(120):
            random.seed(5000 + i)
            t = evaluate_triggers(s, now)
            assert t is not None, f"seed {i}: 高孤独+must_send 不得空"
            if t.type == "lonely_mid" and t.intensity == "soft":
                seen_downgrade += 1
                assert t.data.get("must_send") is True, \
                    f"seed {i}: 降级路径 must_send 丢失: {t.data}"
        assert seen_downgrade > 0, "安全阀降级应至少发生一次"
    print("  OK test_a4_must_send_preserved_across_safety_downgrade")


def test_a6_repeat_damping_same_type_only():
    """A6 统一 repeat 阻尼：同类型在 trigger_history 重复 n 次 → 权重 ×0.6^min(n,3)；
    异类型历史（morning）不影响；cap=3 封顶（6 次 == 3 次）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)

        def run(history, n=300):
            s = _make_state(td, now, loneliness=75, energy=40)
            s.cooldown.trigger_history = list(history)
            return _run_seeds(s, now, n=n)

        base = run([])
        base_low = base.get("lonely_low", 0)
        assert base_low > 80, f"基线 lonely_low 应占优, got {base}"
        # 同类型重复 3 次 → lonely_low ×0.6^3=0.216 → 选中率显著下降
        damped = run(["lonely_low"] * 3)
        damped_low = damped.get("lonely_low", 0)
        assert damped_low < base_low * 0.5, \
            f"lonely_low 应显著衰减: {base_low} → {damped_low}, got {damped}"
        # 异类型历史（morning）→ 与空历史计数一致（互不影响，±15% 容差防 RNG 消费漂移）
        other = run(["morning"] * 3)
        other_low = other.get("lonely_low", 0)
        assert abs(other_low - base_low) <= base_low * 0.15, \
            f"异类型历史不得影响: {base_low} vs {other_low}"
        # cap=3：6 次重复 == 3 次重复（min(n,3) 封顶，±15% 容差）
        capped = run(["lonely_low"] * 6)
        capped_low = capped.get("lonely_low", 0)
        assert abs(capped_low - damped_low) <= damped_low * 0.15, \
            f"cap 应封顶: {damped_low} vs {capped_low}"
    print("  OK test_a6_repeat_damping_same_type_only")
# A5: 未回复退场状态机
# ═══════════════════════════════════════════════════════════

def test_backoff_level_boundaries():
    """三态边界：<3 → 0；3-4 → 1；≥5 → 2；参数可配"""
    from chiguo_trigger import backoff_level
    now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
    with tempfile.TemporaryDirectory() as td:
        # 默认参数 [cooldown].backoff_start=3 / backoff_silent=5
        for n, expect in [(0, 0), (1, 0), (2, 0), (3, 1), (4, 1), (5, 2), (6, 2)]:
            s = _make_state(td, now, energy=40)
            s.cooldown.messages_without_reply = n
            assert backoff_level(s, now) == expect, f"n={n} expect {expect}"
        # 参数覆盖
        s = _make_state(td, now, energy=40)
        s.config["cooldown"]["backoff_start"] = 2
        s.config["cooldown"]["backoff_silent"] = 4
        s.cooldown.messages_without_reply = 2
        assert backoff_level(s, now) == 1
        s.cooldown.messages_without_reply = 4
        assert backoff_level(s, now) == 2
    print("  OK test_backoff_level_boundaries")


def test_backoff_level1_suppresses_emotional_keeps_ritual():
    """backing_off（3-4 条未回复）：情绪类禁发、仪式类照发"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        # 14:00 无仪式候选 + 高孤独：情绪类全被禁 → 全 None
        s = _make_state(td, now, loneliness=75, energy=40, messages_without_reply=3)
        counts = _run_seeds(s, now, n=100)
        assert counts.get("None", 0) == 100, f"backing_off must suppress all emotional, got {counts}"
        # 09:00 早安窗口：仪式类照发（morning ~20/200），情绪类仍为 0
        now2 = datetime(2026, 6, 15, 9, 0, tzinfo=CST)
        s2 = _make_state(td, now2, loneliness=75, energy=40, messages_without_reply=4)
        counts2 = _run_seeds(s2, now2, n=200)
        assert counts2.get("morning", 0) > 0, f"ritual must still fire, got {counts2}"
        emotional = {k: v for k, v in counts2.items()
                     if k in ("lonely_low", "lonely_mid", "lonely_high", "anxiety",
                              "playful", "reflect", "longing")}
        assert not emotional, f"emotional must be suppressed in backing_off, got {counts2}"
    print("  OK test_backoff_level1_suppresses_emotional_keeps_ritual")


def test_backoff_level2_silent_suppresses_all():
    """silent（≥5 条未回复）：全禁发（仪式类也被禁）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 9, 0, tzinfo=CST)  # 早安窗口
        s = _make_state(td, now, loneliness=75, energy=40, messages_without_reply=5)
        counts = _run_seeds(s, now, n=100)
        assert counts.get("None", 0) == 100, f"silent must suppress everything, got {counts}"
    print("  OK test_backoff_level2_silent_suppresses_all")


def test_backoff_level2_escape_valve_exempt():
    """silent 态 escape_valve longing 破防豁免（防死锁语义必须保留）"""
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, anxiety=85, energy=40, messages_without_reply=5)
        # 高焦虑阻塞 + 墙钟沉默 ≥72h + 无冷却记录 → longing_break_eligible
        s.cooldown.last_user_message_at = (now - timedelta(hours=100)).isoformat()
        assert s.longing_break_eligible(now) is True, "precondition: escape valve eligible"
        t = evaluate_triggers(s, now)
        assert t is not None, "escape_valve must bypass silent suppression"
        assert t.type == "longing" and t.data.get("escape_valve") is True, t
    print("  OK test_backoff_level2_escape_valve_exempt")


# ═══════════════════════════════════════════════════════════
# #83 回归: None 防护 / 配置类型防护 / 记忆兜底类型+概率门控
# ═══════════════════════════════════════════════════════════

def test_is_free_time_holiday_parser_none_no_crash():
    """#83 Bug1: HolidayParser 构造失败降级（holiday_parser=None）→
    _is_free_time 不抛 AttributeError 且视为空闲（无假日判定 → 空闲）"""
    from chiguo_trigger import _is_free_time
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        s = _make_state(td, now, energy=40)
        s.holiday_parser = None
        assert _is_free_time(s, now) is True, \
            "holiday_parser=None 时无假日判定 → 应视为空闲"
    print("  OK test_is_free_time_holiday_parser_none_no_crash")


def test_backoff_level_bad_config_falls_back():
    """#83 Bug2: backoff_start/silent 为 "3.5"/None 等非整数 → 不崩溃且回退默认 3/5"""
    from chiguo_trigger import backoff_level
    now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
    with tempfile.TemporaryDirectory() as td:
        # 字符串小数 → 回退默认 3/5：n=4 → level 1
        s = _make_state(td, now, energy=40)
        s.config["cooldown"]["backoff_start"] = "3.5"
        s.config["cooldown"]["backoff_silent"] = "5.7"
        s.cooldown.messages_without_reply = 4
        assert backoff_level(s, now) == 1, '"3.5"/"5.7" 应回退默认 3/5 → level 1'
        # None → 回退默认 3/5：n=5 → level 2
        s2 = _make_state(td, now, energy=40)
        s2.config["cooldown"]["backoff_start"] = None
        s2.config["cooldown"]["backoff_silent"] = None
        s2.cooldown.messages_without_reply = 5
        assert backoff_level(s2, now) == 2, 'None 应回退默认 3/5 → level 2'
        # 合法整数配置不受影响
        s3 = _make_state(td, now, energy=40)
        s3.config["cooldown"]["backoff_start"] = 2
        s3.config["cooldown"]["backoff_silent"] = 4
        s3.cooldown.messages_without_reply = 2
        assert backoff_level(s3, now) == 1, "合法整数配置应正常工作"
    print("  OK test_backoff_level_bad_config_falls_back")


class _FakeMemoryBridge:
    """memory_bridge 替身：available=True，user_relevant 返回预设记忆（含调用计数）"""
    available = True

    def __init__(self, mems=None):
        self.mems = list(mems or [])
        self.calls = 0

    def user_relevant(self, limit=10, min_importance=0.4):
        self.calls += 1
        return list(self.mems)

    def random_memory(self, min_importance=0.4):
        return None


def test_memory_fallback_string_timestamp_no_crash():
    """#83 Bug3①: 记忆兜底 timestamp 为 ISO 字符串 → 跳过该条不崩溃，
    后续正常 epoch 条目仍可处理（防 str > float TypeError）"""
    from unittest import mock
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        epoch_ts = int(now.timestamp() - 5 * 3600)  # 5h 前 → 48h 窗口内
        bridge = _FakeMemoryBridge([
            {"timestamp": "2026-06-15T10:00:00", "text": "字符串时间戳"},
            {"timestamp": epoch_ts, "text": "正常 epoch 记忆"},
        ])
        s = _make_state(td, now, energy=40)
        s.memory_bridge = bridge
        with mock.patch("chiguo_trigger.random.random", return_value=0.0):
            t = evaluate_triggers(s, now)  # 不得抛异常
        assert bridge.calls == 1, "门控放行时应执行一次兜底搜索"
        assert t is None or t.type == "follow_up", f"兜底命中应为 follow_up, got {t}"
    print("  OK test_memory_fallback_string_timestamp_no_crash")


def test_memory_fallback_probability_gate():
    """#83 Bug3②: 记忆兜底块必须有概率门控（pending_topics 为空时
    不得每 tick 无条件多关键词搜索）——random 恒 1.0 拦截、恒 0.0 放行"""
    from unittest import mock
    with tempfile.TemporaryDirectory() as td:
        now = datetime(2026, 6, 15, 14, 0, tzinfo=CST)
        bridge = _FakeMemoryBridge([])
        s = _make_state(td, now, energy=40)
        s.memory_bridge = bridge
        with mock.patch("chiguo_trigger.random.random", return_value=1.0):
            evaluate_triggers(s, now)
        assert bridge.calls == 0, "概率门控应拦截兜底搜索"
        with mock.patch("chiguo_trigger.random.random", return_value=0.0):
            evaluate_triggers(s, now)
        assert bridge.calls == 1, "概率门控放行时应执行兜底搜索"
    print("  OK test_memory_fallback_probability_gate")


# ═══════════════════════════════════════════════════════════
# 入口
# ═══════════════════════════════════════════════════════════
