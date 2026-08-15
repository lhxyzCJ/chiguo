#!/usr/bin/env python3
"""test_schedule_parser.py — schedule 模块单元测试（解析/缓存刷新）"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import date, datetime, timezone, timedelta
from schedule.parsing import parse_cell, parse_weeks

CST = timezone(timedelta(hours=8))

def dt(y, m, d, hh, mm):
    return datetime(y, m, d, hh, mm, tzinfo=CST)

# ── 解析层:周数 ──

def test_parse_weeks_single():
    assert parse_weeks("19周") == {19}
    print("  OK test_parse_weeks_single")

def test_parse_weeks_range():
    assert parse_weeks("2-17周") == set(range(2, 18))
    print("  OK test_parse_weeks_range")

def test_parse_weeks_odd_even():
    assert parse_weeks("10-16(双)周") == {10, 12, 14, 16}
    assert parse_weeks("3-15单") == {3, 5, 7, 9, 11, 13, 15}
    print("  OK test_parse_weeks_odd_even")

def test_parse_weeks_comma():
    assert parse_weeks("2-4,6,8-10周") == {2, 3, 4, 6, 8, 9, 10}
    print("  OK test_parse_weeks_comma")

def test_parse_weeks_garbage():
    assert parse_weeks("abc") == set()
    print("  OK test_parse_weeks_garbage")

# ── 解析层:单元格 ──

def test_parse_cell_single_hyphen():
    r = parse_cell("高等数学BII(理论)-刘洋【2-17周】尚行楼")
    assert len(r) == 1 and r[0]["course"] == "高等数学BII(理论)"
    assert r[0]["teacher"] == "刘洋"
    assert r[0]["weeks"] == set(range(2, 18))
    assert r[0]["location"] == "尚行楼"
    print("  OK test_parse_cell_single_hyphen")

def test_parse_cell_single_space():
    r = parse_cell("管理学基础(理论) 王芳【2-17周】 南楼610智慧教室")
    assert len(r) == 1 and r[0]["course"] == "管理学基础(理论)"
    assert r[0]["teacher"] == "王芳"
    print("  OK test_parse_cell_single_space")

def test_parse_cell_merged_two_courses():
    cell = "工程CAD实训-张伟【19周】尚行楼304 BIM实训室4  管理学基础(理论) 王芳【2-17周】 南楼610智慧教室"
    r = parse_cell(cell)
    assert len(r) == 2, r
    assert r[0]["course"] == "工程CAD实训" and r[1]["course"] == "管理学基础(理论)"
    print("  OK test_parse_cell_merged_two_courses")

def test_parse_cell_trailing_location_fragment():
    # 尾部纯 location 残片（location 内含 2+ 空白）→ 回退整 cell 解析
    r = parse_cell("工程CAD实训-张伟【19周】尚行楼304  BIM实训室4")
    assert len(r) == 1 and r[0]["course"] == "工程CAD实训", r
    print("  OK test_parse_cell_trailing_location_fragment")

def test_parse_cell_truncated_course():
    # 失败段含【 → 残缺课程丢弃
    r = parse_cell("高等数学-刘洋【2-17周】尚行楼  残缺课")
    assert len(r) == 1 and r[0]["course"] == "高等数学", r
    print("  OK test_parse_cell_truncated_course")

def test_parse_cell_empty():
    assert parse_cell("") == []
    assert parse_cell("  ") == []
    print("  OK test_parse_cell_empty")

# ── xlsx 解析 + 缓存（refresh_schedule_cache，真实 openpyxl fixture）──

import tempfile, json, time
from pathlib import Path
import openpyxl

from schedule.parser import refresh_schedule_cache
from schedule.query import current_period


def test_current_period_boundaries():
    assert current_period(dt(2026, 6, 15, 8, 0)) == 1
    assert current_period(dt(2026, 6, 15, 8, 47)) is None  # 课间
    assert current_period(dt(2026, 6, 15, 20, 40)) == 11
    assert current_period(dt(2026, 6, 15, 22, 0)) is None  # 晚课后
    print("  OK test_current_period_boundaries")


def _write_xlsx(path: Path, rows: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, row in enumerate(rows, start=5):  # 第5行起是课表数据
        ws.cell(row=i, column=1, value=row[0])
        for j, cell in enumerate(row[1:], start=2):
            ws.cell(row=i, column=j, value=cell)
    wb.save(str(path))


ROWS = [
    [1, "高等数学BII(理论)-刘洋【2-17周】尚行楼", None, None, None, None, None, None],
    [3, None, None, "管理学基础(理论) 王芳【2-17周】 南楼610智慧教室", None, None, None, None],
    [4, "工程CAD实训-张伟【19周】尚行楼304 BIM实训室4  管理学基础(理论) 王芳【2-17周】 南楼610智慧教室", None, None, None, None, None, None],
]


def test_refresh_creates_cache():
    """首次 refresh：解析 xlsx → 落盘缓存"""
    with tempfile.TemporaryDirectory() as td:
        xp = Path(td) / "xskb.xlsx"
        _write_xlsx(xp, ROWS)
        cp = Path(td) / "c.json"
        ok = refresh_schedule_cache(str(xp), str(cp), date(2026, 2, 23))
        assert ok is True
        assert cp.exists()
        data = json.loads(cp.read_text())
        assert data["cache_version"] == 2
        # 周一第1节（col 2 = weekday 0）与合并单元格备选落入缓存
        assert data["schedule"]["0"]["1"]["course"] == "高等数学BII(理论)", data
        assert data["schedule"]["0"]["4"]["alternates"], data
        print("  OK test_refresh_creates_cache")


def test_refresh_reparses_on_mtime():
    """xlsx mtime 变化 → 重新解析；未变 → 复用缓存不重写"""
    with tempfile.TemporaryDirectory() as td:
        xp = Path(td) / "xskb.xlsx"
        _write_xlsx(xp, ROWS)
        cp = Path(td) / "c.json"
        assert refresh_schedule_cache(str(xp), str(cp), date(2026, 2, 23))
        parsed_at_1 = json.loads(cp.read_text())["parsed_at"]
        assert refresh_schedule_cache(str(xp), str(cp), date(2026, 2, 23))
        assert json.loads(cp.read_text())["parsed_at"] == parsed_at_1
        time.sleep(1.1)
        _write_xlsx(xp, [[1, "新课程-老师【2-17周】新地点", None, None, None, None, None, None]])
        assert refresh_schedule_cache(str(xp), str(cp), date(2026, 2, 23))
        data = json.loads(cp.read_text())
        assert data["parsed_at"] > parsed_at_1
        assert data["schedule"]["0"]["1"]["course"] == "新课程", data
        print("  OK test_refresh_reparses_on_mtime")


def test_refresh_cache_corrupt_is_replaced():
    """坏缓存 → 删除并重新解析落盘"""
    with tempfile.TemporaryDirectory() as td:
        xp = Path(td) / "xskb.xlsx"
        _write_xlsx(xp, ROWS)
        cp = Path(td) / "c.json"
        cp.write_text("{corrupt json")
        assert refresh_schedule_cache(str(xp), str(cp), date(2026, 2, 23))
        assert json.loads(cp.read_text())["cache_version"] == 2
        print("  OK test_refresh_cache_corrupt_is_replaced")


def test_refresh_parse_failure_keeps_old_cache():
    """xlsx 损坏 → 解析失败：保留旧缓存文件，返回仍 True（有可用缓存）"""
    with tempfile.TemporaryDirectory() as td:
        xp = Path(td) / "xskb.xlsx"
        _write_xlsx(xp, ROWS)
        cp = Path(td) / "c.json"
        assert refresh_schedule_cache(str(xp), str(cp), date(2026, 2, 23))
        xp.write_text("not an xlsx")
        time.sleep(1.1)
        os.utime(xp)
        ok = refresh_schedule_cache(str(xp), str(cp), date(2026, 2, 23))
        assert ok is True, "旧缓存仍可用 → available=True"
        assert cp.exists(), "失败解析不应覆盖旧缓存文件"
        print("  OK test_refresh_parse_failure_keeps_old_cache")


def test_refresh_v1_migration_forces_reparse():
    """v1 缓存 → 强制重解析"""
    with tempfile.TemporaryDirectory() as td:
        xp = Path(td) / "xskb.xlsx"
        _write_xlsx(xp, ROWS)
        cp = Path(td) / "c.json"
        cp.write_text(json.dumps({"cache_version": 1, "parsed_at": 0, "schedule": {}}))
        assert refresh_schedule_cache(str(xp), str(cp), date(2026, 2, 23))
        data = json.loads(cp.read_text())
        assert "0" in data["schedule"], "v1 缓存强制重解析"
        print("  OK test_refresh_v1_migration_forces_reparse")


def test_refresh_disabled():
    """enabled=False → 不解析不落盘"""
    with tempfile.TemporaryDirectory() as td:
        cp = Path(td) / "c.json"
        ok = refresh_schedule_cache(str(Path(td) / "none.xlsx"), str(cp), date(2026, 2, 23), enabled=False)
        assert ok is False
        assert not cp.exists()
        print("  OK test_refresh_disabled")



