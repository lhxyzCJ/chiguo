# schedule/parser.py — 课表缓存刷新（数据面）
# 解析 xskb.xlsx → schedule_cache.json（daemon 启动时刷新；mtime 变化自动重新解析）
# 零 token 消耗，确定性解析。读路径 = schedule.sources.load_sources 直读缓存。
# 旧版 ScheduleParser 类的 .query() 生产零调用（读路径不走它），故收敛为纯刷新函数。

import json
import sys
from datetime import date
from pathlib import Path

from schedule.parsing import parse_cell
from chiguo_atomic import atomic_write  # Q23: 共享原子写助手


def refresh_schedule_cache(xlsx_path: str, cache_path: str, semester_start: date,
                           enabled: bool = True) -> bool:
    """xlsx 变更时重新解析并落盘 schedule_cache.json。

    返回 available：False 表示无可用课表数据（enabled=False / xlsx 缺失 / 解析失败）。
    解析失败（xlsx 损坏/openpyxl 缺失等）：保留旧缓存，不覆盖落盘。"""
    xp = Path(xlsx_path)
    cp = Path(cache_path)
    if not enabled:
        return False

    schedule, parsed_at = _load_cache(cp)
    if not xp.exists():
        # xlsx 缺失时保留缓存课表（缓存不存在则保持空课表）
        return bool(schedule)
    xlsx_mtime = xp.stat().st_mtime
    if xlsx_mtime <= parsed_at:
        return True  # 缓存新鲜，直接复用
    parsed = _parse(xp)
    if parsed:
        _save_cache(cp, parsed, xlsx_mtime)
        return True
    return bool(schedule)


def _parse(xp: Path) -> dict:
    """解析 xlsx，提取每节课信息。失败（openpyxl 缺失/文件损坏/空表）返回空 dict。"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(xp), read_only=True, data_only=True)
    except Exception as e:
        print(f"[schedule_parser] xlsx parse failed ({e}), schedule=empty", file=sys.stderr)
        return {}

    # Issue #403：read_only workbook 持有 fd + 临时解压目录，必须显式 close；
    # try/finally 保证解析成功/失败/中途异常都不泄漏（此前靠 GC 回收）。
    try:
        if not wb.sheetnames:
            raise ValueError("xlsx has no sheets")
        ws = wb[wb.sheetnames[0]]

        schedule: dict = {}
        for row in ws.iter_rows(min_row=5, max_row=15, values_only=True):  # 第5行起是课表数据
            if not row or not row[0]:
                continue
            try:
                period = int(float(str(row[0])))
            except (ValueError, TypeError):
                continue
            if period < 1 or period > 11:
                continue

            for col_idx in range(1, 8):  # 周一~周日
                if col_idx >= len(row):
                    continue
                cell = str(row[col_idx]).strip() if row[col_idx] else ""
                if not cell or cell == "None":
                    continue

                courses = parse_cell(cell)
                if not courses:
                    continue

                weekday = col_idx - 1  # 0=Mon ... 6=Sun
                if weekday not in schedule:
                    schedule[weekday] = {}
                if len(courses) == 1:
                    schedule[weekday][period] = courses[0]
                else:
                    # 合并单元格：同一时段多门课（周次互斥，如 2-17 与 19 周）。
                    # 主课程存原字段，其余存入 alternates
                    schedule[weekday][period] = {
                        **courses[0], "alternates": courses[1:]
                    }
        return schedule
    except Exception as e:
        print(f"[schedule_parser] xlsx parse failed ({e}), schedule=empty", file=sys.stderr)
        return {}
    finally:
        wb.close()


def _save_cache(cp: Path, schedule: dict, parsed_at: float) -> None:
    def make_serializable(c):
        c = {**c, "weeks": sorted(c["weeks"])}
        c["alternates"] = [
            {**a, "weeks": sorted(a["weeks"])}
            for a in (c.get("alternates") or [])
        ]
        return c

    data = {
        "cache_version": 2,
        "parsed_at": parsed_at,
        "schedule": {
            str(day): {
                str(period): make_serializable(course)
                for period, course in periods.items()
            }
            for day, periods in schedule.items()
        }
    }
    atomic_write(cp, json.dumps(data, indent=2, ensure_ascii=False), mode=0o600)


def _load_cache(cp: Path) -> tuple[dict, float]:
    """读缓存。损坏 → 删除坏文件并忽略（避免 daemon 崩溃）。"""
    if not cp.exists():
        return {}, 0
    try:
        data = json.loads(cp.read_text())
        if not isinstance(data, dict):
            raise ValueError("cache root must be a dict")
        parsed_at = data.get("parsed_at", 0)
        schedule = {}
        for day_str, periods in data.get("schedule", {}).items():
            day = int(day_str)
            schedule[day] = {}
            for period_str, course in periods.items():
                c = dict(course)
                c["weeks"] = set(c.get("weeks", []))  # list → set
                # 合并单元格的备选课程（旧缓存无此字段 → 空列表）
                c["alternates"] = [
                    {**a, "weeks": set(a.get("weeks", []))}
                    for a in (c.get("alternates") or [])
                ]
                schedule[day][int(period_str)] = c
    except (json.JSONDecodeError, ValueError, TypeError, AttributeError, OSError):
        try:
            cp.unlink()
        except OSError:
            pass
        return {}, 0
    # 旧版本缓存（合并单元格课被吞进 location）→ 强制重解析（xlsx 存在时）
    if data.get("cache_version", 1) < 2:
        parsed_at = 0
    return schedule, parsed_at
