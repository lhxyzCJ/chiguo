# ============================================================
# schedule_parser.py — 课表解析器
# 解析 xskb.xlsx → schedule_cache.json → query(now) → 上课状态
# 零 token 消耗，确定性解析，mtime 变化自动重新解析
# ============================================================

import json
import os
import re
import sys
from datetime import datetime, date, timedelta, timezone
from pathlib import Path

CST = timezone(timedelta(hours=8))

# ── 节次 → 时间映射（中国大学标准） ──────────────────────

PERIOD_TIMES = {
    1:  ("08:00", "08:45"),
    2:  ("08:50", "09:35"),
    3:  ("10:00", "10:45"),
    4:  ("10:50", "11:35"),
    5:  ("14:00", "14:45"),
    6:  ("14:50", "15:35"),
    7:  ("16:00", "16:45"),
    8:  ("16:50", "17:35"),
    9:  ("19:00", "19:45"),
    10: ("19:50", "20:35"),
    11: ("20:40", "21:25"),
}


class ScheduleParser:
    """课表解析 + 查询"""

    def __init__(self, xlsx_path: str = "data/xskb.xlsx",
                 cache_path: str = "schedule_cache.json",
                 semester_start: date = None,
                 enabled: bool = True):
        self.xlsx_path = Path(xlsx_path)
        self.cache_path = Path(cache_path)
        self.enabled = enabled          # 可选来源开关（false → 完全不解析，query 返回空课表）
        self.available = False          # 课表数据是否可用（enabled 且有解析/缓存）
        # 学期起始日期，默认 2026-02-23（可配置）
        self.semester_start = semester_start or date(2026, 2, 23)
        self._schedule: dict = {}   # {weekday: {period: course_info}}
        self._parsed_at: float = 0
        self._ensure_parsed()

    # ── 解析 ──────────────────────────────────────────────

    def _ensure_parsed(self):
        """如果 xlsx 被更新，重新解析"""
        if not self.enabled:
            return
        self._load_cache()
        if not self.xlsx_path.exists():
            # xlsx 缺失时保留缓存课表（缓存不存在则保持空课表）
            return
        xlsx_mtime = self.xlsx_path.stat().st_mtime
        if xlsx_mtime <= self._parsed_at:
            return  # 缓存新鲜，直接复用
        if self._parse():
            self._parsed_at = xlsx_mtime
            self.available = True
            self._save_cache()
        else:
            # 解析失败（xlsx 损坏/openpyxl 缺失等）：降级空课表，但保留旧缓存
            # （不用空数据覆盖落盘）。记住 mtime 避免每轮 query 重复解析刷 stderr，
            # xlsx 修复后 mtime 变化会自然重试。
            self._parsed_at = xlsx_mtime

    def _parse(self) -> bool:
        """解析 xlsx，提取每节课信息。失败（openpyxl 缺失/文件损坏/空表）降级空课表，不崩溃。
        返回 bool：True = 解析成功（调用方应落盘缓存）；False = 失败（保留旧缓存）。"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(self.xlsx_path))
            if not wb.sheetnames:
                raise ValueError("xlsx has no sheets")
            ws = wb[wb.sheetnames[0]]
        except Exception as e:
            print(f"[schedule_parser] xlsx parse failed ({e}), schedule=empty", file=sys.stderr)
            self._schedule = {}
            return False

        self._schedule = {}
        weekday_map = {i: i - 1 for i in range(1, 8)}  # col 2→Mon(0), col 8→Sun(6)

        for row in ws.iter_rows(min_row=5, max_row=15, values_only=True):  # 第5行起是课表数据
            if not row or not row[0]:
                continue
            try:
                period = int(float(str(row[0])))
            except (ValueError, TypeError):
                continue
            if period < 1 or period > 11:
                continue

            for col_idx in range(1, 8):  # 周一~周日
                if col_idx >= len(row):
                    continue
                cell = str(row[col_idx]).strip() if row[col_idx] else ""
                if not cell or cell == "None":
                    continue

                courses = self._parse_cell(cell)
                if not courses:
                    continue

                weekday = col_idx - 1  # 0=Mon ... 6=Sun
                if weekday not in self._schedule:
                    self._schedule[weekday] = {}
                if len(courses) == 1:
                    self._schedule[weekday][period] = courses[0]
                else:
                    # 合并单元格：同一时段多门课（周次互斥，如 2-17 与 19 周）。
                    # 主课程存原字段，其余存入 alternates
                    self._schedule[weekday][period] = {
                        **courses[0], "alternates": courses[1:]
                    }
        return True

    # 课程名-教师 分隔符：`-`（工程CAD实训-张伟）、`- `（工程测量实训- 李娜）、
    # 空格（管理学基础(理论) 王芳）。教师名限 2-4 个汉字，其后紧跟【周数】。
    _COURSE_PART_RE = re.compile(
        r'^(.+?)[- ]+([\u4e00-\u9fa5]{2,4})\s*【(.+?)】(.*)$'
    )
    # 旧正则（课程名-教师必须为横杠分隔），仅作为回退
    _LEGACY_COURSE_RE = re.compile(r'^(.+?)-(.+?)【(.+?)】(.*)$')
    # 合并单元格把多门课拼成一个 cell，课程间以 2+ 连续空白分隔
    _CELL_SPLIT_RE = re.compile(r'\s{2,}')

    @staticmethod
    def _parse_cell(cell: str) -> list[dict]:
        """
        解析单元格文本，返回课程列表（合并单元格可能含多门课）。
        格式: "课程名-教师【周数】教室" 或 "课程名 教师【周数】教室"
        示例: "高等数学BII(理论)-刘洋【2-17周】尚行楼"
              "工程CAD实训-张伟【19周】尚行楼304 BIM实训室4  管理学基础(理论) 王芳【2-17周】 南楼610智慧教室"
        """
        cell = cell.replace("\r", " ").replace("\n", " ").strip()
        if not cell:
            return []

        # 合并单元格：按 2+ 连续空白拆成多段，每段是一门完整课程
        parts = [p for p in ScheduleParser._CELL_SPLIT_RE.split(cell) if p]
        courses = []
        for part in parts:
            c = ScheduleParser._parse_course_part(part)
            if not c:
                c = ScheduleParser._parse_course_legacy(part)
            if c:
                courses.append(c)

        if len(parts) <= 1:
            return courses

        # 多段场景：
        # 1) 最后一段解析成功 → 正常的多课程合并，全部保留
        last = ScheduleParser._parse_course_part(parts[-1])
        if not last:
            last = ScheduleParser._parse_course_legacy(parts[-1])
        if last:
            return courses

        # 2) 失败段里含 【 → 尾部是残缺课程，丢弃（避免被吞进 location）
        if any("【" in p for p in parts[len(courses):]):
            return courses

        # 3) 尾部是纯 location 残片（如 location 内部恰有 2+ 空白导致误拆）：
        #    回退为整 cell 解析（新正则优先，保留完整课程；location 略脏可接受）
        c = ScheduleParser._parse_course_part(cell)
        if not c:
            c = ScheduleParser._parse_course_legacy(cell)
        return [c] if c else courses

    @staticmethod
    def _parse_course_part(part: str) -> dict | None:
        """解析单段课程文本：课程名[- ]教师【周数】地点"""
        match = ScheduleParser._COURSE_PART_RE.match(part)
        if not match:
            return None
        return ScheduleParser._make_course(*match.groups())

    @staticmethod
    def _parse_course_legacy(part: str) -> dict | None:
        """旧格式回退：课程名-教师【周数】地点"""
        match = ScheduleParser._LEGACY_COURSE_RE.match(part)
        if not match:
            return None
        return ScheduleParser._make_course(*match.groups())

    @staticmethod
    def _make_course(course_name: str, teacher: str, weeks_str: str,
                     location: str) -> dict:
        """构造课程 dict（解析周数模式）"""
        return {
            "course": course_name.strip(),
            "teacher": teacher.strip(),
            "weeks": ScheduleParser._parse_weeks(weeks_str),  # set of week numbers
            "weeks_raw": weeks_str.strip(),
            "location": location.strip(),
        }

    @staticmethod
    def _parse_weeks(weeks_str: str) -> set[int]:
        """
        解析周数表达式 → 周数集合。
        "19周" → {19}
        "2-17周" → {2,3,...,17}
        "10-16(双)周" → {10,12,14,16}
        "2-4,6,8-10周" → {2,3,4,6,8,9,10}
        """
        weeks = set()
        # 移除 "周" 字
        s = weeks_str.replace("周", "").replace(" ", "")

        # 处理单/双周标记：括号形式 (单)/(双) 或后缀形式 3-15单 / 3-15双
        odd_even = None
        if "(单)" in s:
            odd_even = "odd"
            s = s.replace("(单)", "")
        elif "(双)" in s:
            odd_even = "even"
            s = s.replace("(双)", "")
        elif odd_even is None and "单" in s:
            odd_even = "odd"
            s = s.replace("单", "")
        elif odd_even is None and "双" in s:
            odd_even = "even"
            s = s.replace("双", "")

        # 按逗号分割
        for part in s.split(","):
            part = part.strip()
            if not part:
                continue
            if "-" in part:
                lo, hi = part.split("-", 1)
                try:
                    lo, hi = int(lo), int(hi)
                    for w in range(lo, hi + 1):
                        if odd_even == "odd" and w % 2 == 0:
                            continue
                        if odd_even == "even" and w % 2 != 0:
                            continue
                        weeks.add(w)
                except ValueError:
                    pass
            else:
                try:
                    weeks.add(int(part))
                except ValueError:
                    pass
        return weeks

    # ── 缓存 ──────────────────────────────────────────────

    def _save_cache(self):
        def make_serializable(c):
            c = {**c, "weeks": sorted(c["weeks"])}
            c["alternates"] = [
                {**a, "weeks": sorted(a["weeks"])}
                for a in (c.get("alternates") or [])
            ]
            return c

        data = {
            "cache_version": 2,
            "parsed_at": self._parsed_at,
            "schedule": {
                str(day): {
                    str(period): make_serializable(course)
                    for period, course in periods.items()
                }
                for day, periods in self._schedule.items()
            }
        }
        tmp_path = Path(str(self.cache_path) + ".tmp")
        tmp_path.write_text(json.dumps(data, indent=2, ensure_ascii=False))
        os.replace(tmp_path, self.cache_path)

    def _load_cache(self):
        if not self.cache_path.exists():
            return
        try:
            data = json.loads(self.cache_path.read_text())
            if not isinstance(data, dict):
                raise ValueError("cache root must be a dict")
            parsed_at = data.get("parsed_at", 0)
            schedule = {}
            for day_str, periods in data.get("schedule", {}).items():
                day = int(day_str)
                schedule[day] = {}
                for period_str, course in periods.items():
                    c = dict(course)
                    c["weeks"] = set(c.get("weeks", []))  # list → set
                    # v2: 合并单元格的备选课程（旧缓存无此字段 → 空列表）
                    c["alternates"] = [
                        {**a, "weeks": set(a.get("weeks", []))}
                        for a in (c.get("alternates") or [])
                    ]
                    schedule[day][int(period_str)] = c
        except (json.JSONDecodeError, ValueError, TypeError, AttributeError, OSError):
            # 缓存损坏：删除坏文件并忽略，避免 daemon 崩溃
            try:
                self.cache_path.unlink()
            except OSError:
                pass
            return
        self._parsed_at = parsed_at
        self._schedule = schedule
        self.available = True
        # v2: 旧版本缓存（合并单元格课被吞进 location）→ 强制重解析（xlsx 存在时）
        if data.get("cache_version", 1) < 2:
            self._parsed_at = 0

    # ── 查询 ──────────────────────────────────────────────

    def query(self, now: datetime) -> dict:
        """
        查询当前上课状态。
        返回:
          {
            "in_class": bool,           # 是否在上课
            "current_course": dict|null, # 当前课程信息
            "next_course": dict|null,   # 下一节课
            "periods_today": [...],     # 今天所有课
            "class_load": "heavy"|"normal"|"light"|"free",
          }
        """
        self._ensure_parsed()

        weekday = now.weekday()  # 0=Mon ... 6=Sun
        current_period = self._current_period(now)
        # 学期第几周（从 semester_start 算起）
        week_num = max(1, (now.date() - self.semester_start).days // 7 + 1)

        today_courses = self._schedule.get(weekday, {})

        # 筛选本周有效的课程。
        # 合并单元格（alternates）同一时段多门课周次互斥：取 weeks 含本周的那门
        active_courses = {}
        for period, entry in today_courses.items():
            for course in [entry] + entry.get("alternates", []):
                weeks = course.get("weeks", set())
                if weeks and week_num in weeks:
                    active_courses[period] = course
                    break

        # 当前课程
        in_class = False
        current_course = None
        if current_period in active_courses:
            in_class = True
            current_course = {k: v for k, v in active_courses[current_period].items()
                              if k != "alternates"}
            current_course["period"] = current_period
            current_course["time"] = PERIOD_TIMES.get(current_period, ("?", "?"))
            # 距离下课剩余分钟数
            end_time_str = PERIOD_TIMES.get(current_period, ("?", "?"))[1]
            end_h, end_m = map(int, end_time_str.split(":"))
            end_time = now.replace(hour=end_h, minute=end_m, second=0)
            current_course["minutes_remaining"] = max(0, (end_time - now).total_seconds() / 60)

        # 下一节课
        next_course = None
        next_free_at = None
        sorted_periods = sorted(active_courses.keys())
        for p in sorted_periods:
            if p > (current_period or 0):
                next_course = {k: v for k, v in active_courses[p].items()
                               if k != "alternates"}
                next_course["period"] = p
                next_course["time"] = PERIOD_TIMES.get(p, ("?", "?"))
                break
        if next_course is None and in_class:
            # 下课后就自由了
            next_free_at = PERIOD_TIMES.get(current_period, ("?", "?"))[1]

        # 课业负担
        total_periods = len(active_courses)
        remaining = len([p for p in sorted_periods if p > (current_period or 0)])
        if total_periods == 0:
            class_load = "free"
        elif total_periods <= 2:
            class_load = "light"
        elif total_periods <= 5:
            class_load = "normal"
        else:
            class_load = "heavy"

        return {
            "in_class": in_class,
            "current_course": current_course,
            "next_course": next_course,
            "next_free_at": next_free_at,
            "periods_today": [
                {"period": p, **{k: v for k, v in active_courses[p].items()
                                 if k != "alternates"}}
                for p in sorted(active_courses.keys())
            ],
            "class_load": class_load,
            "remaining_classes": remaining,
            "total_classes": total_periods,
            "available": self.available,
        }

    @staticmethod
    def _current_period(now: datetime) -> int | None:
        """根据当前时间返回所在节次"""
        current_minutes = now.hour * 60 + now.minute
        for period, (start_str, end_str) in PERIOD_TIMES.items():
            sh, sm = map(int, start_str.split(":"))
            eh, em = map(int, end_str.split(":"))
            start = sh * 60 + sm
            end = eh * 60 + em
            if start <= current_minutes <= end:
                return period

            # 课间：如果在两节课之间，返回前一节
            next_start = None
            if period + 1 in PERIOD_TIMES:
                ns = PERIOD_TIMES[period + 1][0]
                nsh, nsm = map(int, ns.split(":"))
                next_start = nsh * 60 + nsm

            if end < current_minutes and (next_start is None or current_minutes < next_start):
                # 在课间休息中
                return None  # 课间视为不在上课

            if period == max(PERIOD_TIMES.keys()) and current_minutes > end:
                return None

        return None


# ── CLI ────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    from datetime import datetime, timezone, timedelta
    CST = timezone(timedelta(hours=8))

    parser = ScheduleParser("data/xskb.xlsx")

    if len(sys.argv) > 1 and sys.argv[1] == "--dump":
        print(json.dumps(parser._schedule, indent=2, ensure_ascii=False, default=str))
    else:
        now = datetime.now(CST)
        result = parser.query(now)
        print(json.dumps(result, indent=2, ensure_ascii=False, default=str))
